from openpyxl import load_workbook
import random
wb = load_workbook("Single Machine Scheduling.xlsx")
ws = wb.active

ready_time = []
row_title = ws[2][1:]
for cell in row_title:
    ready_time.append(cell.value)

processing_time = []
row_title = ws[3][1:]
for cell in row_title:
    processing_time.append(cell.value)

due_date = []
row_title = ws[4][1:]
for cell in row_title:
    due_date.append(cell.value)

weight = []
row_title = ws[5][1:]
for cell in row_title:
    weight.append(cell.value)

schedule = [1,2,3]

def calculate_makespan(ready_time, processing_time, schedule):

    makespan = 0
    completion_time_list = []

    for i in range(len(schedule)):
        job_index = schedule[i] - 1
        if i == 0 :
            completion_time = ready_time[job_index] + processing_time[job_index]
            completion_time_list.append(completion_time)
        else:
            if completion_time_list[i - 1] < ready_time[job_index]:
                completion_time = ready_time[job_index] + processing_time[job_index]
            else:
                completion_time = completion_time_list[i - 1] + processing_time[job_index]
            completion_time_list.append(completion_time)

    makespan = completion_time_list[len(schedule) - 1]
    return makespan

def calculate_total_flowtime(ready_time, processing_time, schedule):

    flowtime = 0
    completion_time_list = []

    for i in range(len(schedule)):
        job_index = schedule[i] - 1
        if i == 0 :
            completion_time = ready_time[job_index] + processing_time[job_index]
            completion_time_list.append(completion_time)
        else:
            if completion_time_list[i - 1] < ready_time[job_index]:
                completion_time = ready_time[job_index] + processing_time[job_index]
            else:
                completion_time = completion_time_list[i - 1] + processing_time[job_index]
            completion_time_list.append(completion_time)

    flowtime = sum(completion_time_list)
    return flowtime

def calculate_total_tardiness(ready_time, processing_time, due_date, schedule):

    total_tardiness = 0
    completion_time_list = []
    total_tardiness_list = []

    for i in range(len(schedule)):
        job_index = schedule[i] - 1

        if i == 0:
            completion_time = ready_time[job_index] + processing_time[job_index]
            completion_time_list.append(completion_time)
            tardiness = completion_time_list[i] - due_date[job_index]
            if tardiness <= 0 :
                tardiness = 0
            else:
                tardiness = completion_time_list[i] - due_date[job_index]
            total_tardiness_list.append(tardiness)
        else:
            if completion_time_list[i - 1] <= ready_time[job_index]:
                completion_time = ready_time[job_index] + processing_time[job_index]
                completion_time_list.append(completion_time)
                tardiness = completion_time_list[i] - due_date[job_index]
                if tardiness <= 0:
                    tardiness = 0
                else:
                    tardiness = completion_time_list[i] - due_date[job_index]
                total_tardiness_list.append(tardiness)
            else:
                completion_time = completion_time_list[i - 1] + processing_time[job_index]
                completion_time_list.append(completion_time)
                tardiness = completion_time_list[i] - due_date[job_index]
                if tardiness <= 0:
                    tardiness = 0
                else:
                    tardiness = completion_time_list[i] - due_date[job_index]
                total_tardiness_list.append(tardiness)
        total_tardiness += total_tardiness_list[i]
    return total_tardiness
    
def calculate_total_weighted_tardiness(ready_time, processing_time, due_date, weight, schedule):

    total_weight_tardiness = 0
    completion_time_list = []
    total_tardiness_list = []

    for i in range(len(schedule)):
        job_index = schedule[i] - 1

        if i == 0:
            completion_time = ready_time[job_index] + processing_time[job_index]
            completion_time_list.append(completion_time)
            tardiness = completion_time_list[i] - due_date[job_index]
            if tardiness <= 0:
                tardiness = 0
            else:
                tardiness = completion_time_list[i] - due_date[job_index]
            total_tardiness_list.append(tardiness)
        else:
            if completion_time_list[i - 1] <= ready_time[job_index]:
                completion_time = ready_time[job_index] + processing_time[job_index]
                completion_time_list.append(completion_time)
                tardiness = completion_time_list[i] - due_date[job_index]
                if tardiness <= 0:
                    tardiness = 0
                else:
                    tardiness = completion_time_list[i] - due_date[job_index]
                total_tardiness_list.append(tardiness)
            else:
                completion_time = completion_time_list[i - 1] + processing_time[job_index]
                completion_time_list.append(completion_time)
                tardiness = completion_time_list[i] - due_date[job_index]
                if tardiness <= 0:
                    tardiness = 0
                else:
                    tardiness = completion_time_list[i] - due_date[job_index]
                total_tardiness_list.append(tardiness)
        total_weight_tardiness += total_tardiness_list[i] * weight[job_index]
    return total_weight_tardiness


print("SCHEDULE: ", str(schedule))
print("makespan값:",calculate_makespan(ready_time, processing_time, schedule))
print("total flowtime값:",calculate_total_flowtime(ready_time, processing_time, schedule))
print("total tardiness값:",calculate_total_tardiness(ready_time, processing_time, due_date, schedule))
print("total weight tardiness값:",calculate_total_weighted_tardiness(ready_time, processing_time, due_date, weight, schedule))

# print("makespan값:",calculate_makespan(ready_time, processing_time, schedule))
# ws["A8"] = "MAKESPAN"
# ws["B8"] = calculate_makespan(ready_time, processing_time, schedule)
# ws["A9"] = "과제순서"

#wb.save("Single Machine Scheduling2.xlsx")