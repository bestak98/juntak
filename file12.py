import random
import openpyxl  # Openpyxl 라이브러리 불러오기
import numpy as np
import pandas as pd
from copy import deepcopy as cp


operation_num_dict = {"A": 3, "B": 4}
PRODUCTION_REQUIREMENT = {'A': 10, 'B': 20}
machine_type_num = 2
not_available_rate = 0.2
SETUP_TIME = {'homogeneous_setup': 2, 'heterogeneous_setup': 4}


#파일1- 잘 변하지 않는 정보
def first_file (operation_num_dict,PRODUCTION_REQUIREMENT,machine_type_num,not_available_rate,SETUP_TIME):
    global wb1
    wb1 = openpyxl.Workbook()  # Workbook 만들기
    new_filename = './FJS1.xlsx'  # 파일 위치 및 이름 정하기

    # 파일 1 시트1,2,3,4 생성
    wb1_ws = wb1.active # 가장 처음 만들어진 Sheet를 변수로 선언
    wb1_ws.title = "Operation Type" # 첫 Sheet의 이름 정하기
    wb1_ws2 = wb1.create_sheet("Machine Type", 1)
    wb1_ws3 = wb1.create_sheet("Type Processing Time", 2)
    wb1_ws4 = wb1.create_sheet("Setup Time", 3)

    #파일1 시트1 열생성
    wb1_ws.cell(row=1, column=1).value = "Job Index"
    wb1_ws.cell(row=1, column=2).value = "Job Type"
    wb1_ws.cell(row=1, column=3).value = "Operation Type"

    #파일1 시트1 행생성, 시트3 행생성
    p = 2
    job_type_list = list(operation_num_dict.keys())
    n = len(job_type_list)
    global total_operation_list
    total_operation_list = []
    for i in range(n):
        wb1_ws.cell(row=i+2, column=1).value = i #index 번호
        wb1_ws.cell(row=i+2, column=2).value = job_type_list[i]  # 엑셀은 셀이 (1,1)부터 시작
    for i in range(n):
        empty_list = []
        for m in range(operation_num_dict[list(operation_num_dict.keys())[i]]):
            job_operation = list(operation_num_dict)[i]+"_"+str(m+1)
            total_operation_list.append(job_operation)
            empty_list.append(job_operation)
            wb1_ws3.cell(row=p, column=1).value = job_operation
            p = p + 1
        # print(str(empty_list))
        oper_string = ""
        for j in empty_list:
            oper_string += str(j) + " "
        oper_string = oper_string.strip()
        # print(oper_string)
        # print(oper_string.strip().split(" "))

        wb1_ws.cell(row=i+2, column=3).value = oper_string

    #파일1 시트2

    #파일1 시트2 행생성
    for i in range(1, machine_type_num+1):
        wb1_ws2.cell(row=i, column=1).value = "Type" + str(i)

    #파일1 시트2 열생성
    wb1_ws2.cell(row=1, column=2).value = "DA"
    wb1_ws2.cell(row=2, column=2).value = "WB"

    #파일1 시트3

    #파일1 시트3 열생성
    for i in range(machine_type_num):
        wb1_ws3.cell(row=1, column=i+2).value = "Type" + str(i+1)


    #파일1 시트3 값입력
    global operation_num
    operation_num = 0
    for values in operation_num_dict.values():
        
        operation_num += values
    
    value_num = machine_type_num*operation_num

    sig_sigma, sig_mu = 1, 2   # 표준편차, 평균 = 2, 2
    sig_list = sig_sigma * np.random.randn(1, value_num) + sig_mu
    sig_list = abs(sig_list).tolist()
    sig_list = [int (i) for i in sig_list[0]]

    mu_sigma, mu_mu = 2, 10
    mu_list = mu_sigma * np.random.randn(1, value_num) + mu_mu
    mu_list = abs(mu_list).tolist()
    mu_list = [int (i) for i in mu_list[0]]


    value_list = []
    for i in range(value_num):
        value_list.append(int(random.normalvariate(mu_list[i], sig_list[i])))

    k = 0
    for i in range(machine_type_num):
        for j in range(operation_num):
            wb1_ws3.cell(j+2, i+2).value = int(value_list[k])
            k += 1

    # 해당 machine에서 operation이 수행될 수 없는 경우

    # operation이 수행될 수 없는 경우의 개수 구하기 - 반올림
    total_value_num = operation_num * machine_type_num
    not_available_num = round(total_value_num * not_available_rate)

    # 안되는 경우를 랜덤으로 지정하기
    rand_lst = [] # 랜덤으로 지정된 행과 열을 저장하는 리스트
    available_lst = [] # 불가능한 셀을 저장하는 리스트
    n = 0
    while True:
        row_num = random.randint(1, operation_num)
        column_num = random.randint(1, machine_type_num)
        if (row_num, column_num) in rand_lst:
            continue
        elif (row_num, column_num) in available_lst:
            continue
        else:
            wb1_ws3.cell(row_num+1, column_num+1).value = "X"
            n += 1
        rand_lst.append((row_num, column_num))
        for j in range(1, machine_type_num+1):
            if j != column_num:
                available_lst.append((row_num, j))
        if n == not_available_num:
            break


    #파일1 시트4
    setup_type_list = list(SETUP_TIME.keys())
    setup_time_list = list(SETUP_TIME.values())

    for i in range(len(setup_type_list)):
        wb1_ws4.cell(i + 1, 1).value = setup_type_list[i]
    for j in range(len(setup_time_list)):
        wb1_ws4.cell(j + 1, 2).value = setup_time_list[j]

    wb1.save(new_filename)  # 파일 저장하기
    return 

first_file(operation_num_dict,PRODUCTION_REQUIREMENT,machine_type_num,not_available_rate,SETUP_TIME)

#파일2- 자주 변하는 정보

def second_file(wb1,operation_num,total_operation_list,PRODUCTION_REQUIREMENT,machine_type_num):
    wb2 = openpyxl.Workbook()  # Workbook 만들기
    new_filename = './FJS2.xlsx'  # 파일 위치 및 이름 정하기


    # 파일2 시트1,2,3 생성
    wb2_ws = wb2.active # 가장 처음 만들어진 Sheet를 변수로 선언
    wb2_ws.title = "Production Requirement" # 첫 Sheet의 이름 정하기
    wb2_ws2 = wb2.create_sheet("Machine number", 1)
    wb2_ws3 = wb2.create_sheet("Setup Status", 2)



    #파일 2 시트1

    #파일2 시트1 행이름생성 - job_type
    i = 1
    for keys in PRODUCTION_REQUIREMENT.keys():
        wb2_ws.cell(row=i, column=1).value = keys
        i += 1

    #파일2 시트1 값생성 - production_num
    job_type_num = len(PRODUCTION_REQUIREMENT)
    for i in range(1, job_type_num+1):
        wb2_ws.cell(i, 2).value = list(PRODUCTION_REQUIREMENT.values())[i-1]


    #파일2 시트2

    #파일2 시트2 행이름생성
    for i in range(2):
        wb2_ws2.cell(i+1,1).value = wb1["Machine Type"].cell(i+1,2).value

    #파일2 시트2 Machine Type별 개수 입력
    sigma, mu = 2, 10   # 표준편차, 평균 = 2, 10
    value = sigma * np.random.randn(machine_type_num,1) + mu
    machine_lot_front = 1
    machine_lot_behind_list=[]
    for i in range(len(value)):
        for j in range(len(value[0])):
            machine_lot_behind = machine_lot_front + int(value[i][j])
            wb2_ws2.cell(i+1, j+2).value = str(machine_lot_front) + "~" + str(machine_lot_behind)
            machine_lot_behind_list.append(machine_lot_behind)
            machine_lot_front = machine_lot_behind + 1
            
            
    #파일2 시트3

    #파일2 시트3 양식 생성
    for i in range(machine_lot_front - 1):
        wb2_ws3.cell(i+2,1).value = "M_" + str(i+1)

    wb2_ws3.cell(1,2).value = "Setup Status"

    #파일2 시트3 Setup Status값 입력

    total_processing_time_list = []
    for i in range(machine_type_num):
        for j in range(operation_num):
            total_processing_time_list.append(wb1["Type Processing Time"].cell(j+2,i+2).value)
    #print(available_operation_list)

    for j in range(machine_type_num):
        no_oper_list = []
        for n in range(len(total_processing_time_list[j*operation_num:(j+1)*operation_num])):
            if total_processing_time_list[j*operation_num:(j+1)*operation_num][n] == "X":
                no_oper_list.append(total_operation_list[n]) #실행 불가능한 o을 모으는 리스트            
        #print(no_oper_list)
        available_operation_list = list(set(total_operation_list)-set(no_oper_list)) #전체 o_list에서 실행 불가능한 o를 뺀 차집합    
                        
        for i in range(machine_lot_front - 1):            
            type_machine_num = wb2_ws2.cell(j+1,2).value
            if i+1 in range(int(type_machine_num.split("~")[0]),int(type_machine_num.split("~")[1])+1):
                #print(available_operation_list)
                
                wb2_ws3.cell(i+2,2).value = random.choice(available_operation_list) #차집합 중 랜덤으로 요소 선택
                

    wb2.save(new_filename)  # 파일 저장하기

second_file(wb1,operation_num,total_operation_list,PRODUCTION_REQUIREMENT,machine_type_num)