from copy import deepcopy as cp
import pandas as pd
import openpyxl

wb1 = openpyxl.load_workbook(filename = 'FJS1.xlsx')
wb2 = openpyxl.load_workbook(filename = 'FJS2.xlsx')


df = pd.read_excel('./FJS2.xlsx', sheet_name = 'Production Requirement',header=None, index_col=0)
PRODUCTION_REQUIREMENT = df.to_dict()[1]
# PRODUCTION_REQUIREMENT = {'A' : 10, 'B' : 20}
print(PRODUCTION_REQUIREMENT)

operation_num_dict = {}
for i in range(len(PRODUCTION_REQUIREMENT)):
    # print(i, wb1['Operation Type'].cell(i+2,2).value )
    operation_num_dict[wb1['Operation Type'].cell(i+2,2).value] = wb1['Operation Type'].cell(i+2,3).value[-1]
# print(operation_num_dict) #operation_num_dict = {"A": 3, "B": 4}

operation_list = [] #진행해야할 모든 operation , 추후에 정해진 rule에 따라 SCHEDULE's KEY(Machine Number)들의 VALUE로 들어갈 것
a=0
for key,value in dict(PRODUCTION_REQUIREMENT).items():
    for i in range(a,value):
        for n in range(int(operation_num_dict[key])):
            operation_list.append('LOT_'+str(i+1)+'-'+key+'_'+str(n+1))
    a = value
print('-OPERATION LIST-\n',operation_list)
    

SCHEDULE = {'M_1': ['LOT_1-A_1', 'LOT_1-A_2', 'LOT_2-B_1'],
         'M_2': ['LOT_1-A_3'],
         'M_3': ['LOT_3-B_1', 'LOT_3-B_2', 'LOT_2-B_2'],
         'M_4': ['LOT_4-C_2', 'LOT_4-C_1']}

df = pd.read_excel('./FJS1.xlsx', sheet_name = 'Type Processing Time', header=0, index_col=0)
file_read = df.to_dict()
PROCESSING_TIME = {}
for key, value in file_read.items():
    for op_key, time_value in value.items():
        PROCESSING_TIME[key+'-'+op_key] = time_value

print('-PROCESSING_TIME-\n',PROCESSING_TIME)
# PROCESSING_TIME = { 'M_1-A_1': 4, 'M_1-A_2': 4, 'M_1-B_1': 5, 'M_2-A_3': 1, 'M_3-B_1': 4, 'M_3-B_2': 2, 'M_4-C_1': 3, 'M_4-C_2': 3}

df = pd.read_excel('./FJS2.xlsx', sheet_name = 'Setup Status',header=0, index_col=0)
INITIAL_SETUP = df.to_dict()['Setup Status']
print('-INITIAL_SETUP-\n',INITIAL_SETUP)

#INITIAL_SETUP = {'M_1' : 'A_1', 'M_2' : 'A_1', 'M_3' : 'A_1', 'M_4' : 'A_1'}

df = pd.read_excel('./FJS1.xlsx', sheet_name = 'Operation Type',header=0)
JOBTYPE_JOB = dict([(i,a) for i,a in zip(df['Job Type'],df['Job Index'])])
# JOBTYPE_JOB = {'A': 0, 'B': 1, 'C': 2}

df = pd.read_excel('./FJS1.xlsx', sheet_name = 'Setup Time', header=None, index_col=0)
SETUP_TIME = df.to_dict()[1]
# SETUP_TIME = {'homogeneous_setup': 2, 'heterogeneous_setup': 4}
# SETUP_TIME # homogeneous setup ex) 1_1 -> 1_3 : 2 / heterogeneous ex) 1_1 -> 2_2 : 4

num_of_lot = 0
for value in PRODUCTION_REQUIREMENT.values():
    num_of_lot += value
num_of_job_type = len(PRODUCTION_REQUIREMENT.keys())
num_of_machine = len(SCHEDULE.keys())


job_completion_time_list = []
machine_completion_time_list = []
job_current_finished_operation_index_list = []
completion_list = []
latest_completion_dict = {}

for i in range(num_of_job_type):
    job_completion_time_list.append(0)
for i in range(num_of_machine):
    machine_completion_time_list.append(0)

remain_SCHEDULE_dict = cp(SCHEDULE)

#job별로 현재까지 수행한 operation
for i in range(num_of_lot):
    job_current_finished_operation_index_list.append(0)


remain_operation_lst = SCHEDULE.values() # machine 구분 있음 #초기 리스트

# 중간에 있는 operation의 시작 시간 = max(같은 job의 이전 operation의 완료시간, 같은 기계의 직전 operation의 완료시간)
while True:
    is_break = True
    # remain_num = 0  # machine별 가장 처음 operation이 1로 끝나는게 있는지 확인하는 변수
    # # 딕셔너리에 남은 OPERATION이 없을 때 코드 종료
    # # for values in remain_SCHEDULE_dict.values():
    # #     if not values:
    # #         remain_num += 1
    # # if remain_num == num_of_machine:
    # #     break

    for key, value in remain_SCHEDULE_dict.items():
        if value:
            operation = value[0]
            lot_num = int(operation.split('-')[0].split('_')[1])
            if int(operation.split('_')[2]) - 1 == job_current_finished_operation_index_list[lot_num-1]:
                is_break = False
                completion_list.append(operation)
                processing_time = PROCESSING_TIME[key + '-' + operation.split('-')[1]]
                max_value = max(job_completion_time_list[JOBTYPE_JOB[operation.split('-')[1][0]]], machine_completion_time_list[int(key.split('_')[1]) - 1])
                job_completion_time_list[JOBTYPE_JOB[operation.split('-')[1][0]]] = max_value + PROCESSING_TIME[key + '-' + operation.split('-')[1]]  # 끝나는 시간 = 시작시간 + 생산시간
                machine_completion_time_list[int(key.split('_')[1]) - 1] = max_value + PROCESSING_TIME[key + '-' + operation.split('-')[1]]
                job_current_finished_operation_index_list[lot_num - 1] += 1

                if key not in latest_completion_dict.keys():
                    setup = INITIAL_SETUP[key]
                else:
                    setup = latest_completion_dict[key]

                operation_cut = operation.split('-')[1]
                latest_completion_dict[key] = operation_cut
                setup_time = 0
                if setup.split('_')[0] == operation_cut.split('_')[0]:
                    if setup.split('_')[1] == operation_cut.split('_')[1]:
                        pass
                    else:
                        setup_time = SETUP_TIME['homogeneous_setup']
                        job_completion_time_list[JOBTYPE_JOB[operation_cut.split('_')[0]]] += SETUP_TIME['homogeneous_setup']
                        machine_completion_time_list[int(key.split('_')[1]) - 1] += SETUP_TIME['homogeneous_setup']
                else:
                    setup_time = SETUP_TIME['heterogeneous_setup']
                    job_completion_time_list[JOBTYPE_JOB[operation_cut.split('_')[0]]] += SETUP_TIME['heterogeneous_setup']
                    machine_completion_time_list[int(key.split('_')[1]) - 1] += SETUP_TIME['heterogeneous_setup']
                print("SETUP", key, operation.split("-")[1], max_value, max_value + setup_time, )
                print("PROCESSING", key, operation.split("-")[1], max_value + setup_time, max_value + setup_time + processing_time)
                del (remain_SCHEDULE_dict[key][0])
                print(remain_SCHEDULE_dict)
                break
    if is_break:
        break

# 오류
for i in range(num_of_machine):
    if len(list(remain_SCHEDULE_dict.values())[i]) != 0:
        print("Error: Check SCHEDULE")
        exit()
# dict(Machine="M-1", Operation ='A-1', Start='2021-01-20 00', Finish='2021-01-20 05',Job_Type ='Job Type A'),
#     dict(Machine="M-2", Operation ='B-1', Start='2021-01-20 03', Finish='2021-01-20 07',Job_Type ='Job Type B'),
#     dict(Machine="M-1", Operation ='B-2', Start='2021-01-20 11', Finish='2021-01-20 14',Job_Type ='Job Type B'),
#     dict(Machine="M-1", Operation ='Setup',Start='2021-01-20 07', Finish='2021-01-20 11',Job_Type ='Setup Time'),
#     dict(Machine="M-2", Operation ='Setup',Start='2021-01-20 07', Finish='2021-01-20 11',Job_Type ='Setup Time'),
#     dict(Machine="M-2", Operation='A-1', Start='2021-01-20 11', Finish='2021-01-20 19', Job_Type='Job Type A'),
print(machine_completion_time_list)
print(job_completion_time_list)
print(max(job_completion_time_list))