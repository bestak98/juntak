from copy import deepcopy as cp
import numpy as np
import openpyxl

wb = openpyxl.Workbook()  # Workbook 만들기
new_filename = './Gantt.xlsx'  # 파일 위치 및 이름 정하기
ws = wb.active # 가장 처음 만들어진 Sheet를 변수로 선언

PRODUCTION_REQUIREMENT = {'A' : 1, 'B' : 2, 'C' : 1} # A-3, B-2, C-2
SCHEDULE = {'M_1': ['LOT_1-A_1', 'LOT_1-A_2', 'LOT_2-B_1'],
         'M_2': ['LOT_1-A_3'],
         'M_3': ['LOT_3-B_1', 'LOT_3-B_2', 'LOT_2-B_2'],
         'M_4': ['LOT_4-C_1', 'LOT_4-C_2']}


PROCESSING_TIME = { 'M_1-A_1': 4, 'M_1-A_2': 4, 'M_1-B_1': 5, 'M_2-A_3': 1, 'M_3-B_1': 4, 'M_3-B_2': 2, 'M_4-C_1': 3, 'M_4-C_2': 3}
# SETUP_TIME # homogeneous setup ex) 1_1 -> 1_3 : 2 / heterogeneous ex) 1_1 -> 2_2 : 4
INITIAL_SETUP = {'M_1' : 'A_1', 'M_2' : 'A_1', 'M_3' : 'A_1', 'M_4' : 'A_1'}
JOBTYPE_JOB = {'A': 0, 'B': 1, 'C': 2}

SETUP_TIME = {'homogeneous_setup': 2, 'heterogeneous_setup': 4}

num_of_lot = 4
num_of_job_type = 3
num_of_machine = 4

job_completion_time_list = []
machine_completion_time_list = []
job_current_finished_operation_index_list = []
completion_list = []
array_list = []
setup_array_list = []
latest_completion_dict = {}

for i in range(num_of_job_type):
    job_completion_time_list.append(0)
for i in range(num_of_machine):
    machine_completion_time_list.append(0)

remain_operation_dict = cp(PROCESSING_TIME)
remain_SCHEDULE_dict = cp(SCHEDULE)

#job별로 현재까지 수행한 operation
for i in range(num_of_lot):
    job_current_finished_operation_index_list.append(0)


remain_operation_lst = SCHEDULE.values() # machine 구분 있음 #초기 리스트

# 중간에 있는 operation의 시작 시간 = max(같은 job의 이전 operation의 완료시간, 같은 기계의 직전 operation의 완료시간)
while True:
    remain_num = 0  # machine별 가장 처음 operation이 1로 끝나는게 있는지 확인하는 변수
    # 딕셔너리에 남은 OPERATION이 없을 때 코드 종료
    for values in remain_SCHEDULE_dict.values():
        if not values:
            remain_num += 1
    if remain_num == num_of_machine:
        break

    for key, value in remain_SCHEDULE_dict.items():
        if value:
            operation = value[0]
            lot_num = int(operation.split('-')[0].split('_')[1])
            if int(operation.split('_')[2]) - 1 == job_current_finished_operation_index_list[lot_num-1]:
                completion_list.append(operation)
                max_value = max(job_completion_time_list[JOBTYPE_JOB[operation.split('-')[1][0]]], machine_completion_time_list[int(key.split('_')[1]) - 1])
                job_completion_time_list[JOBTYPE_JOB[operation.split('-')[1][0]]] = max_value + PROCESSING_TIME[key + '-' + operation.split('-')[1]]  # 끝나는 시간 = 시작시간 + 생산시간
                machine_completion_time_list[int(key.split('_')[1]) - 1] = max_value + PROCESSING_TIME[key + '-' + operation.split('-')[1]]
                job_current_finished_operation_index_list[lot_num - 1] += 1

                array1 = np.array(machine_completion_time_list)
                array_list.append(array1)
                print(array_list)


                if key not in latest_completion_dict.keys():
                    setup = INITIAL_SETUP[key]
                else:
                    setup = latest_completion_dict[key]

                operation_cut = operation.split('-')[1]
                latest_completion_dict[key] = operation_cut

                if setup.split('_')[0] == operation_cut.split('_')[0]:
                    if setup.split('_')[1] == operation_cut.split('_')[1]:
                        pass
                    else:
                        job_completion_time_list[JOBTYPE_JOB[operation_cut.split('_')[0]]] += SETUP_TIME['homogeneous_setup']
                        machine_completion_time_list[int(key.split('_')[1]) - 1] += SETUP_TIME['homogeneous_setup']
                else:
                    job_completion_time_list[JOBTYPE_JOB[operation_cut.split('_')[0]]] += SETUP_TIME['heterogeneous_setup']
                    machine_completion_time_list[int(key.split('_')[1]) - 1] += SETUP_TIME['heterogeneous_setup']

                array2 = np.array(machine_completion_time_list)

                setup_array = array2 - array1
                setup_array_list.append(setup_array)
                print(setup_array_list)

                del (remain_SCHEDULE_dict[key][0])
                break

time_dict = {}

for j in range(num_of_machine):
    time_list = []
    for i in range(len(array_list)):
        setup_time = setup_array_list[i][j]
        if i == 0:
            oper_time = array_list[i][j] - 0
        else:
            oper_time = (array_list[i] - array_list[i-1])[j]
        if oper_time != 0:
            if i == 0:
                time_list.append([0, setup_time])
                time_list.append([setup_time, oper_time])
            else:
                time_list.append([array_list[i][j]-setup_time, array_list[i][j]])
                time_list.append([array_list[i][j], array_list[i][j]+oper_time])
    key = 'M-'+str(j+1)
    time_dict[key] = time_list
print(time_dict)

#열생성
key_list = list(time_dict.keys())
for i in range(1, num_of_machine+1):
    ws.cell(1, i).value = key_list[i-1]
#행생성
    value_list = list(time_dict.values())[i-1]
    for j in range(len(value_list)):
        oper_string = ""
        for k in value_list[j]:
            oper_string += str(k) + " "
        oper_string = oper_string.strip()
        ws.cell(j+2, i).value = oper_string
wb.save(new_filename)  # 파일 저장하기


print(machine_completion_time_list)
print(job_completion_time_list)
print(max(job_completion_time_list))