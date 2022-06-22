import random
import openpyxl  # Openpyxl 라이브러리 불러오기
import numpy as np
import pandas as pd

operation_num_dict = {"A": 3, "B": 4}
machine_type_num = 2
not_available_rate = 0.2
SETUP_TIME = {'homogeneous_setup': 2, 'heterogeneous_setup': 4}

wb1 = openpyxl.Workbook()  # Workbook 만들기
new_filename = './FJS.xlsx'  # 파일 위치 및 이름 정하기

# 시트1,2,3,4 생성
wb1_ws = wb1.active # 가장 처음 만들어진 Sheet를 변수로 선언
wb1_ws.title = "Operation Type" # 첫 Sheet의 이름 정하기
wb1_ws2 = wb1.create_sheet("Machine Type", 1)
wb1_ws3 = wb1.create_sheet("Type Processing Time", 2)
wb1_ws4 = wb1.create_sheet("Setup Time", 3)

#시트1 열생성
wb1_ws.cell(row=1, column=1).value = "Job Index"
wb1_ws.cell(row=1, column=2).value = "Job Type"
wb1_ws.cell(row=1, column=3).value = "Operation Type"

#시트1 행생성, 시트3 행생성
p = 2
job_type_list = list(operation_num_dict.keys())
n = len(job_type_list)
for i in range(n):
    wb1_ws.cell(row=i+2, column=1).value = i
    wb1_ws.cell(row=i+2, column=2).value = job_type_list[i]  # 엑셀은 셀이 (1,1)부터 시작
for i in range(n):
    empty_list = []
    for m in range(operation_num_dict[list(operation_num_dict.keys())[i]]):
        job_operation = list(operation_num_dict)[i]+"-"+str(m+1)

        empty_list.append(job_operation)
        wb1_ws3.cell(row=p, column=1).value = job_operation
        p = p + 1
    # print(str(empty_list))
    oper_string = ""
    for j in empty_list:
        oper_string += str(j) + " "
    oper_string = oper_string.strip()
    # print(oper_string)
    # print(oper_string.strip().split(" "))

    wb1_ws.cell(row=i+2, column=3).value = oper_string

#시트2

#시트2 행생성
for i in range(1, machine_type_num+1):
    wb1_ws2.cell(row=i, column=1).value = "Type" + str(i)

#시트2 열생성
wb1_ws2.cell(row=1, column=2).value = "DA"
wb1_ws2.cell(row=2, column=2).value = "WB"

#시트3

#시트3 열생성
for i in range(1, machine_type_num+1):
    wb1_ws3.cell(row=1, column=i+1).value = "Type" + str(i)

#시트3 행생성


# 값입력
operation_num = 0
for values in operation_num_dict.values():
    operation_num += values

value_num = machine_type_num*operation_num

sig_sigma, sig_mu = 1, 2   # 표준편차, 평균 = 2, 2
sig_list = sig_sigma * np.random.randn(1, value_num) + sig_mu
sig_list = abs(sig_list).tolist()
sig_list = [int (i) for i in sig_list[0]]

mu_sigma, mu_mu = 2, 10
mu_list = mu_sigma * np.random.randn(1, value_num) + mu_mu
mu_list = abs(mu_list).tolist()
mu_list = [int (i) for i in mu_list[0]]


value_list = []
for i in range(value_num):
    value_list.append(int(random.normalvariate(mu_list[i], sig_list[i])))

k = 0
for i in range(machine_type_num):
    for j in range(operation_num):
        wb1_ws3.cell(j+2, i+2).value = int(value_list[k])
        k += 1

# 해당 machine에서 operation이 수행될 수 없는 경우

# operation이 수행될 수 없는 경우의 개수 구하기 - 반올림
total_value_num = operation_num * machine_type_num
not_available_num = round(total_value_num * not_available_rate)

# 안되는 경우를 랜덤으로 지정하기
rand_lst = [] # 랜덤으로 지정된 행과 열을 저장하는 리스트
for i in range(not_available_num):
    row_num = random.randint(1, operation_num)
    column_num = random.randint(1, machine_type_num)
    for k in range(1, machine_type_num):
        if (row_num, column_num) in rand_lst and wb1_ws3.cell(row_num+1, k+1).value == "X":
            continue
        else:
            wb1_ws3.cell(row_num+1, column_num+1).value = "X"
    rand_lst.append((row_num, column_num))

#시트4
setup_type_list = list(SETUP_TIME.keys())
setup_time_list = list(SETUP_TIME.values())

for i in range(len(setup_type_list)):
    wb1_ws4.cell(i + 1, 1).value = setup_type_list[i]
for j in range(len(setup_time_list)):
    wb1_ws4.cell(j + 1, 2).value = setup_time_list[j]

wb1.save(new_filename)  # 파일 저장하기