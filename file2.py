import random
import openpyxl  # Openpyxl 라이브러리 불러오기
import numpy as np
import pandas as pd
from copy import deepcopy as cp

operation_num_dict = {"A": 3, "B": 4}
PRODUCTION_REQUIREMENT = {'A': 10, 'B': 20}
machine_type_num = 2
unperformable_rate = 0.2

wb2 = openpyxl.load_workbook('./FJS.xlsx')

wb = openpyxl.Workbook()  # Workbook 만들기
new_filename = './FJS1.xlsx'  # 파일 위치 및 이름 정하기


# 시트1,2,3 생성
ws = wb.active # 가장 처음 만들어진 Sheet를 변수로 선언
ws.title = "Production Requirement" # 첫 Sheet의 이름 정하기
ws2 = wb.create_sheet("Machine number", 1)
ws3 = wb.create_sheet("Setup Status", 2)

#파일1 합쳤을 때 지우기
p = 2
job_type_list = list(operation_num_dict.keys())
n = len(job_type_list)
total_operation_list = []
for i in range(n):
    ws.cell(row=i+2, column=1).value = job_type_list[i]  # 엑셀은 셀이 (1,1)부터 시작
for i in range(n):
    empty_list = []
    for m in range(operation_num_dict[list(operation_num_dict.keys())[i]]):
        job_operation = list(operation_num_dict)[i]+"-"+str(m+1)
        empty_list.append(job_operation)
        total_operation_list.append(job_operation)
        ws3.cell(row=p, column=1).value = job_operation
        p = p + 1
    ws.cell(row=i+2, column=2).value = str(empty_list)
operation_num = 0
for values in operation_num_dict.values():
    operation_num += values


#시트1

#시트1 행이름생성 - job_type
i = 1
for keys in PRODUCTION_REQUIREMENT.keys():
    ws.cell(row=i, column=1).value = keys
    i += 1

#시트1 값생성 - production_num
job_type_num = len(PRODUCTION_REQUIREMENT)
for i in range(1, job_type_num+1):
    ws.cell(i, 2).value = list(PRODUCTION_REQUIREMENT.values())[i-1]


#시트2

#시트2 행이름생성
for i in range(2):
    ws2.cell(i+1,1).value = wb2["Machine Type"].cell(i+1,2).value

#Machine Type별 개수 입력
sigma, mu = 2, 10   # 표준편차, 평균 = 2, 10
value = sigma * np.random.randn(machine_type_num,1) + mu
machine_lot_front = 1
machine_lot_behind_list=[]
for i in range(len(value)):
    for j in range(len(value[0])):
        machine_lot_behind = machine_lot_front + int(value[i][j])
        ws2.cell(i+1, j+2).value = str(machine_lot_front) + "~" + str(machine_lot_behind)
        machine_lot_behind_list.append(machine_lot_behind)
        machine_lot_front = machine_lot_behind + 1
        
        
#시트3

#시트3 양식 생성
for i in range(machine_lot_front - 1):
    ws3.cell(i+2,1).value = "M-" + str(i+1)

ws3.cell(1,2).value = "Setup Status"

#Setup Status값 입력

total_processing_time_list = []
for i in range(machine_type_num):
    for j in range(operation_num):
        total_processing_time_list.append(wb2["Type Processing Time"].cell(j+2,i+2).value)
#print(available_operation_list)

for j in range(machine_type_num):
    no_oper_list = []
    for n in range(len(total_processing_time_list[j*operation_num:(j+1)*operation_num])):
        if total_processing_time_list[j*operation_num:(j+1)*operation_num][n] == "X":
            no_oper_list.append(total_operation_list[n]) #실행 불가능한 o을 모으는 리스트            
    print(no_oper_list)
    available_operation_list = list(set(total_operation_list)-set(no_oper_list)) #전체 o_list에서 실행 불가능한 o를 뺀 차집합    
    print(available_operation_list)                 
    for i in range(machine_lot_front - 1):            
        type_machine_num = ws2.cell(j+1,2).value
        if i+1 in range(int(type_machine_num.split("~")[0]),int(type_machine_num.split("~")[1])+1):
            
            
            ws3.cell(i+2,2).value = random.choice(available_operation_list) #차집합 중 랜덤으로 요소 선택
            

wb.save(new_filename)  # 파일 저장하기